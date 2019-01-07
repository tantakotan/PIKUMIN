# -*- coding: utf-8 -*-

import os
import openpyxl
from copy import copy
from openpyxl.utils import column_index_from_string, get_column_letter
from datetime import datetime


class ExecPslsx:

    def __init__(self, dict_of_module, path_of_module, dict_of_parameter, dict_of_option):
        self.dict_of_module = dict_of_module
        self.path_of_module = path_of_module
        self.dict_of_parameter = dict_of_parameter
        self.dict_of_option = dict_of_option

        self.list_of_keys = []
        self.path_of_outputdir = ''
        self.key_of_templatestr = ''
        self.path_of_tplxlsx = ''

        self.num_of_space = int(self.dict_of_option['option_rowspace'])

    def get_outputpath(self):
        self.key_of_templatestr = list(self.dict_of_parameter.keys())[0]
        self.path_of_outputdir = os.path.join(os.path.dirname(self.path_of_module), 'output', self.key_of_templatestr)
        os.makedirs(self.path_of_outputdir, exist_ok=True)

    def get_modulekeys(self):
        self.list_of_keys = list(self.dict_of_module.keys())

    def get_rowspace(self, num_of_space):
        self.num_of_space = num_of_space

    def get_template(self):

        dest_wb = openpyxl.Workbook()

        # seirisimasyo...
        for listkey in self.list_of_keys:
            dest_wb.create_sheet(title=listkey)
            dest_ws = dest_wb[listkey]

            position_row = 0

            for i2 in self.dict_of_module[listkey]:

                path_of_wb = os.path.join(self.path_of_module, i2[0])
                source_wb = openpyxl.load_workbook(path_of_wb)
                source_ws = source_wb[i2[1]]

                nrow = source_ws.min_row
                xrow = source_ws.max_row

                for i in range(nrow, xrow + 1):
                    source_row = source_ws.row_dimensions[i]
                    if source_row.height:
                        dest_row = dest_ws.row_dimensions[i + position_row]
                        dest_row.height = source_row.height

                ncol = source_ws.min_column
                xcol = source_ws.max_column

                for i in range(ncol, xcol + 1):
                    source_col = source_ws.column_dimensions[get_column_letter(i)]
                    if source_col.width:
                        dest_col = dest_ws.column_dimensions[get_column_letter(i)]
                        dest_col.width = source_col.width

                for key, x in enumerate(source_ws.iter_rows(min_row=nrow, max_row=xrow)):

                    for ind, source_cell in enumerate(x):
                        dest_cell = dest_ws[source_cell.coordinate]

                        num_of_row = dest_cell.row + position_row
                        num_of_column = column_index_from_string(dest_cell.column)
                        dest_cell = dest_ws.cell(row=num_of_row, column=num_of_column)

                        dest_cell.value = source_cell.value

                        if source_cell.has_style:
                            dest_cell.font = copy(source_cell.font)
                            dest_cell.border = copy(source_cell.border)
                            dest_cell.fill = copy(source_cell.fill)
                            dest_cell.alignment = copy(source_cell.alignment)
                            dest_cell.number_format = copy(source_cell.number_format)
                            dest_cell.protection = copy(source_cell.protection)

                position_row += xrow - nrow + 1 + self.num_of_space

        time_suffix = datetime.now()
        file_name = self.key_of_templatestr + '_template_' + time_suffix.strftime('%Y%m%d') + '.xlsx'

        dest_wb.remove_sheet(dest_wb.get_sheet_by_name('Sheet'))
        self.path_of_tplxlsx = os.path.join(self.path_of_outputdir, file_name)
        dest_wb.save(self.path_of_tplxlsx)

    def return_tplpath(self):
        return self.path_of_tplxlsx

    def coloring_cells(self, list_of_files):
        for filepath in list_of_files:
            wb = openpyxl.load_workbook(filepath)
            for ws in wb.worksheets:
                for index, ws_row in enumerate(ws.iter_rows()):
                    cells = list(cell for cell in ws_row if cell.value and isinstance(cell.value, str))
                    for cell in cells:
                        if cell.value in 'ケーブルガイド':
                            cell.font = openpyxl.styles.fonts.Font(color='FFFFFF')
                            cell.fill = openpyxl.styles.PatternFill(patternType='solid', start_color='000000', end_color='000000')
                wb.save(filepath)
