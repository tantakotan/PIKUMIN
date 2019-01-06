# -*- coding: utf-8 -*-

from datetime import datetime
from jinja2 import Environment, Template
import os
from openpyxl import load_workbook
import copy

class J2Render:
    def __init__(self, dict_of_parameter):
        self.dict_of_parameter = dict_of_parameter
        self.key_of_templatestring = ''
        self.dict_of_parameter_host = {}
        self.text_of_j2 = ''
        self.path_of_module = ''
        self.render_wb = ''

    def create_file(self, key_of_parameter):

        s = datetime.now()
        s2 = key_of_parameter + '_' + s.strftime('%Y%m%d') + '.conf'
        s3 = os.path.join(self.path_of_module, s2)

        with open(s3, mode='w') as f:
            f.write(self.text_of_j2)

        print('create file successfuly...: ' + s3)

    def create_file_xlsx(self, key_of_parameter):

        s = datetime.now()
        s2 = key_of_parameter + '_' + s.strftime('%Y%m%d') + '.xlsx'
        s3 = os.path.join(self.path_of_module, s2)

        self.render_wb.save(s3)

        print('create file successfuly...: ' + s3)

    def create_outputdir(self, path_of_module):
        self.path_of_module = os.path.join(os.path.dirname(path_of_module), 'output', self.key_of_templatestring)
        os.makedirs(self.path_of_module, exist_ok=True)

    def create_template_key(self):
        self.key_of_templatestring = list(self.dict_of_parameter.keys())[0]

    def create_parameter_host(self, key_of_parameter):
        self.dict_of_parameter_host = dict(self.dict_of_parameter[key_of_parameter])

    def create_template(self, template_string):

        s = datetime.now()
        s2 = self.key_of_templatestring + '_template_' + s.strftime('%Y%m%d') + '.conf'
        s3 = os.path.join(self.path_of_module, s2)

        with open(s3, mode='w') as f:
            f.write(template_string)

        print('create template successfuly...: ' + s3)

    def j2_render(self, text_of_tpl):
        env = Environment()
        env.trim_blocks = True
        env.lstrip_blocks = True

        s = env.from_string(text_of_tpl)
        self.text_of_j2 = s.render(self.dict_of_parameter_host)

    def j2_render_xlsx(self, path_of_tplxlsx):
        env = Environment()
        env.trim_blocks = True
        env.lstrip_blocks = True

        self.render_wb = load_workbook(path_of_tplxlsx)
        for render_ws in self.render_wb.worksheets:
            for index, render_ws_row in enumerate(render_ws.iter_rows()):
                string_cells = list(cell for cell in render_ws_row if cell.value and isinstance(cell.value, str))
                for cell in string_cells:
                    cell.value = Template(cell.value).render(self.dict_of_parameter_host)
