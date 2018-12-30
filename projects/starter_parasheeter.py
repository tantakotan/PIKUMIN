# -*- coding: utf-8 -*-

import os
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string

class starter_parasheeter:
    dict_of_ps = {}
    dict_of_nw = {}
    path_of_ps = ''

    def __init__(self, path_of_ps, dict_of_ps, dict_of_nw):
        self.path_of_ps = path_of_ps
        self.dict_of_ps = dict_of_ps
        self.dict_of_nw = dict_of_nw

    def get_template(self):
        path_of_pslsx = os.path.dirname(self.path_of_ps)

        dict_of_swap = self.dict_of_ps
        dict_of_pslsx = {}

        for i in dict_of_swap:
            list_of_pslsx = [[v, k] for k, v in dict_of_swap[i].items()]
            dict_of_pslsx[i] = list_of_pslsx

        list_of_key = list(dict_of_pslsx.keys())

        to_wb = openpyxl.Workbook()

        for i in list_of_key:
            to_wb.create_sheet(title = i)
            to_ws = to_wb.get_sheet_by_name(i)

            irow = 0

            for i2 in dict_of_pslsx[i]:

                path_of_wb = os.path.join(path_of_pslsx, i2[0])
                wb = openpyxl.load_workbook(path_of_wb)
                ws = wb[i2[1]]

                nrow = ws.min_row
                xrow = ws.max_row

                for i, x in enumerate(ws.rows):
                    for i2, x2 in enumerate(x):
                        to_cell = to_ws[x2.coordinate]
                        to_cell = to_ws.cell(row=to_cell.row + irow, column=column_index_from_string(to_cell.column))
                        to_cell.value = x2.value
                        to_cell._style = copy(x2._style)

                irow += xrow - nrow + 2

            to_wb.save('aaaaaaaaaaaaaa.xlsx')
