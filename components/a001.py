# from jinja2 import Template, Environment, FileSystemLoader
import openpyxl

wb = openpyxl.load_workbook('sample.xlsx')

print(type(wb))
# <class 'openpyxl.workbook.workbook.workbook'>

print(wb.sheetnames)
# ['sheet1', 'sheet2']

sheet = wb['TemplateSample']

print(type(sheet))

cell = sheet['A2']

print(type(cell))

print(cell.value)

cell = sheet.cell(row=2, column=2)

print(type(cell))
print(cell.value)


cell = sheet['B1:C4']
print(cell[0])

def get_value_list(t_2d):
    return([[cell.value for cell in row] for row in t_2d])

l_2d = get_value_list(sheet['A1:C4'])

def get_list_2d(sheet, min_row, max_row, min_col, max_col):
    return get_value_list(sheet.iter_rows(min_row=min_row,
                                          max_row=max_row,
                                          min_col=min_col,
                                          max_col=max_col))

l_2d = get_list_2d(sheet, 2, 3, 2, 3)
l_2d2 = get_list_2d(sheet, 2, 5, 2, 5)

print(type(l_2d))
print(l_2d2)



# env = Environment(loader=FileSystemLoader('.'))
# env.trim_blocks = True
# template = env.get_template('sample.tpl')

# data = {'items': ['Kuro', 'lang', 'Python']}
# disp_text = template.render(data)  # 辞書で指定する
# print(disp_text)

